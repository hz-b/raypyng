"""Tools for inspecting an RML beamline into a single summary table."""

from __future__ import annotations

from pathlib import Path

import pandas as pd

from .rml import RMLFile

NOT_AVAILABLE = "n.a."

TABLE_COLUMNS = [
    "name",
    "type",
    "grazingIncAngle",
    "azimuthalAngle",
    "totalWidth",
    "totalLength",
    "materialCoating1",
    "thicknessCoating1",
    "roughnessCoating1",
    "materialCoating2",
    "thicknessCoating2",
    "roughnessCoating2",
    "materialTopLayer",
    "thicknessTopLayer",
    "roughnessTopLayer",
    "slopeErrorSag",
    "slopeErrorMer",
    "x",
    "y",
    "z",
]

EXPORT_COLUMN_LABELS = {
    "name": "name",
    "type": "type",
    "grazingIncAngle": "grazingIncAngle [deg]",
    "azimuthalAngle": "azimuthalAngle [deg]",
    "totalWidth": "totalWidth [mm]",
    "totalLength": "totalLength [mm]",
    "materialCoating1": "materialCoating1",
    "thicknessCoating1": "thicknessCoating1 [nm]",
    "roughnessCoating1": "roughnessCoating1 [nm]",
    "materialCoating2": "materialCoating2",
    "thicknessCoating2": "thicknessCoating2 [nm]",
    "roughnessCoating2": "roughnessCoating2 [nm]",
    "materialTopLayer": "materialTopLayer",
    "thicknessTopLayer": "thicknessTopLayer [nm]",
    "roughnessTopLayer": "roughnessTopLayer [nm]",
    "slopeErrorSag": "slopeErrorSag [sec]",
    "slopeErrorMer": "slopeErrorMer [sec]",
    "x": "x [mm]",
    "y": "y [mm]",
    "z": "z [mm]",
}


def world_position(element) -> tuple[str, str, str]:
    """Return the world-position (x, y, z) of an RML element as formatted strings."""
    if not hasattr(element, "worldPosition"):
        return NOT_AVAILABLE, NOT_AVAILABLE, NOT_AVAILABLE

    position = element.worldPosition
    return (
        f"{float(position.x.cdata):.2f}",
        f"{float(position.y.cdata):.2f}",
        f"{float(position.z.cdata):.2f}",
    )


def _param_value(element, key: str, enabled_only: bool = False) -> str:
    """Return a param value from an RML element, or ``n.a.`` when unavailable."""
    if not hasattr(element, key):
        return NOT_AVAILABLE

    param = getattr(element, key)
    if enabled_only and param.get_attribute("enabled") != "T":
        return NOT_AVAILABLE

    value = param.cdata.strip()
    return value if value else NOT_AVAILABLE


def _export_table(beamline_table: pd.DataFrame) -> pd.DataFrame:
    """Return a copy of the table with unit-aware export column labels."""
    return beamline_table.rename(columns=EXPORT_COLUMN_LABELS)


def build_tables(rml_path: str | Path) -> pd.DataFrame:
    """Build a single beamline-elements table from an RML file."""
    rml = RMLFile(None, template=str(rml_path))
    rows: list[list[str]] = []

    for element in rml.beamline.children():
        rows.append(
            [
                element.get_attribute("name") or NOT_AVAILABLE,
                element.get_attribute("type") or NOT_AVAILABLE,
                _param_value(element, "grazingIncAngle"),
                _param_value(element, "azimuthalAngle"),
                _param_value(element, "totalWidth"),
                _param_value(element, "totalLength"),
                _param_value(element, "materialCoating1", enabled_only=True),
                _param_value(element, "thicknessCoating1", enabled_only=True),
                _param_value(element, "roughnessCoating1", enabled_only=True),
                _param_value(element, "materialCoating2", enabled_only=True),
                _param_value(element, "thicknessCoating2", enabled_only=True),
                _param_value(element, "roughnessCoating2", enabled_only=True),
                _param_value(element, "materialTopLayer", enabled_only=True),
                _param_value(element, "thicknessTopLayer", enabled_only=True),
                _param_value(element, "roughnessTopLayer", enabled_only=True),
                _param_value(element, "slopeErrorSag"),
                _param_value(element, "slopeErrorMer"),
                *world_position(element),
            ]
        )

    return pd.DataFrame(rows, columns=TABLE_COLUMNS)


def save_tables(beamline_table: pd.DataFrame, tables_dir: str | Path) -> Path:
    """Save the beamline table as a CSV file."""
    tables_dir = Path(tables_dir)
    tables_dir.mkdir(parents=True, exist_ok=True)
    beamline_path = tables_dir / "beamline_elements_table.csv"
    _export_table(beamline_table).to_csv(beamline_path, index=False)
    return beamline_path


def save_tables_xlsx(beamline_table: pd.DataFrame, tables_dir: str | Path) -> Path:
    """Save the beamline table as a styled XLSX file."""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required for save_tables_xlsx. Install it with: pip install openpyxl"
        ) from exc

    tables_dir = Path(tables_dir)
    tables_dir.mkdir(parents=True, exist_ok=True)
    beamline_path = tables_dir / "beamline_elements_table.xlsx"
    _export_table(beamline_table).to_excel(beamline_path, index=False)

    workbook = load_workbook(beamline_path)
    worksheet = workbook.active
    worksheet.title = "Beamline Elements"
    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = False

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    first_col_font = Font(bold=True)
    even_fill = PatternFill(fill_type="solid", fgColor="EAF2F8")
    odd_fill = PatternFill(fill_type="solid", fgColor="FDFEFE")
    border = Border(
        left=Side(style="thin", color="B7C3D0"),
        right=Side(style="thin", color="B7C3D0"),
        top=Side(style="thin", color="B7C3D0"),
        bottom=Side(style="thin", color="B7C3D0"),
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    for row_idx, row in enumerate(
        worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, max_col=worksheet.max_column),
        start=2,
    ):
        fill = even_fill if row_idx % 2 == 0 else odd_fill
        for col_idx, cell in enumerate(row, start=1):
            cell.fill = fill
            cell.border = border
            if col_idx == 1:
                cell.font = first_col_font

    for column_cells in worksheet.columns:
        values = [str(cell.value) if cell.value is not None else "" for cell in column_cells]
        max_len = max(len(value) for value in values)
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(
            max(max_len + 2, 12), 28
        )

    workbook.save(beamline_path)
    return beamline_path
